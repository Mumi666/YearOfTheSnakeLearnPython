import openpyxl
from pathlib import Path, PurePath

src_path = './file/调查问卷'
dist_path = './file/result'

p = Path(src_path)

all_excel = [file for file in p.iterdir() if PurePath(file).match('*.xlsx')]

content = []

for file in all_excel:
    username = file.stem
    wb = openpyxl.load_workbook(file)
    sheet = wb.active
    answer1 = sheet.cell(row=5, column=5).value
    answer2 = sheet.cell(row=11, column=5).value
    temp = [username, answer1, answer2]
    content.append(temp)

print(content)

header_row = ['姓名', '第一题', '第二题']

wb = openpyxl.workbook.Workbook()
sheet = wb.create_sheet(index=0, title='result')

sheet.append(header_row)
for row in content:
    sheet.append(row)

wb.save(filename=dist_path+'/result.xlsx')



