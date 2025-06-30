import openpyxl
from pathlib import Path, PurePath

src_path = './file/工资单/工资单.xlsx'
dist_path = './file/工资单/result'

Path(dist_path).mkdir(parents=True, exist_ok=True)

wb = openpyxl.load_workbook(src_path)
sheet = wb.active

header_row = [cell.value for cell in sheet[1]]

for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=1):
    new_wb = openpyxl.Workbook()
    new_sheet = new_wb.active

    new_sheet.append(header_row)
    new_sheet.append([cell.value for cell in row])

    output_file = Path(dist_path) / f'{row[1].value}.xlsx'
    try:
        new_wb.save(output_file)
    except Exception as e:
        print(f"Error saving file {output_file}: {e}")
